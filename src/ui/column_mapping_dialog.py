from __future__ import annotations

import csv
from pathlib import Path

from PyQt6.QtWidgets import (
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QHeaderView,
    QLabel,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
)

# Dropdown options offered per column.
CATEGORY_OPTIONS = [
    "Skip",
    "NAME",
    "EMAIL",
    "PHONE",
    "LOCATION",
    "ORG",
    "JOBTITLE",
    "DOB",
    "POSTCODE",
    "ID",
    "Freetext",
]

# Auto-suggestion rules: lowercased header → category.
HEADER_SUGGESTIONS: dict[str, str] = {
    "first name": "NAME",
    "firstname": "NAME",
    "last name": "NAME",
    "lastname": "NAME",
    "surname": "NAME",
    "name": "NAME",
    "full name": "NAME",
    "manager": "NAME",
    "contact": "NAME",
    "contact name": "NAME",
    "email": "EMAIL",
    "e-mail": "EMAIL",
    "email address": "EMAIL",
    "manager email": "EMAIL",
    "phone": "PHONE",
    "telephone": "PHONE",
    "tel": "PHONE",
    "mobile": "PHONE",
    "cell": "PHONE",
    "fax": "PHONE",
    "address": "LOCATION",
    "street": "LOCATION",
    "street address": "LOCATION",
    "city": "LOCATION",
    "town": "LOCATION",
    "country": "LOCATION",
    "region": "LOCATION",
    "county": "LOCATION",
    "state": "LOCATION",
    "postcode": "POSTCODE",
    "zip": "POSTCODE",
    "zip code": "POSTCODE",
    "postal code": "POSTCODE",
    "company": "ORG",
    "organisation": "ORG",
    "organization": "ORG",
    "employer": "ORG",
    "firm": "ORG",
    "job title": "JOBTITLE",
    "role": "JOBTITLE",
    "position": "JOBTITLE",
    "title": "JOBTITLE",
    "occupation": "JOBTITLE",
    "dob": "DOB",
    "date of birth": "DOB",
    "birthday": "DOB",
    "birth date": "DOB",
    "birthdate": "DOB",
    "notes": "Freetext",
    "comments": "Freetext",
    "description": "Freetext",
    "bio": "Freetext",
    "remarks": "Freetext",
    "id": "Skip",
    "#": "Skip",
    "row": "Skip",
    "index": "Skip",
}


def _suggest_category(header: str) -> str:
    """Return the best auto-suggested category for a column header."""
    key = header.strip().lower()
    if key in HEADER_SUGGESTIONS:
        return HEADER_SUGGESTIONS[key]
    # Partial matching for compound headers like "Manager Email".
    for pattern, cat in HEADER_SUGGESTIONS.items():
        if pattern in key:
            return cat
    return "Freetext"


def read_csv_headers_and_samples(
    path: Path, sample_rows: int = 3,
) -> tuple[list[str], list[list[str]]]:
    """Read the header row and a few sample rows from a CSV file."""
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        samples: list[list[str]] = []
        for row in reader:
            samples.append(row)
            if len(samples) >= sample_rows:
                break
    return headers, samples


def read_xlsx_headers_and_samples(
    path: Path, sample_rows: int = 3,
) -> tuple[list[str], list[list[str]]]:
    """Read the header row and sample rows from the first XLSX sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True, max_row=sample_rows + 1))
    wb.close()

    if not rows:
        return [], []
    headers = [str(v) if v is not None else "" for v in rows[0]]
    samples = [
        [str(v) if v is not None else "" for v in row]
        for row in rows[1:]
    ]
    return headers, samples


class ColumnMappingDialog(QDialog):
    """Dialog allowing the user to map CSV/XLSX column headers to PII categories."""

    def __init__(
        self,
        headers: list[str],
        samples: list[list[str]],
        parent=None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("Column Mapping")
        self.setMinimumWidth(700)
        self.setMinimumHeight(400)
        self.setModal(True)

        layout = QVBoxLayout(self)
        instructions = QLabel(
            "<b>Map each column to a PII category.</b><br/>"
            "\"Skip\" ignores the column; \"Freetext\" uses automatic detection."
        )
        instructions.setWordWrap(True)
        layout.addWidget(instructions)

        self._headers = headers
        self._combos: list[QComboBox] = []

        table = QTableWidget(len(headers), 3)
        table.setHorizontalHeaderLabels(["Column Header", "Category", "Sample Values"])
        table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents
        )
        table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.ResizeToContents
        )
        table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch
        )

        for row_idx, header in enumerate(headers):
            # Column header (read-only).
            item = QTableWidgetItem(header)
            item.setFlags(item.flags() & ~item.flags().ItemIsEditable)
            table.setItem(row_idx, 0, item)

            # Category combo.
            combo = QComboBox()
            combo.addItems(CATEGORY_OPTIONS)
            suggested = _suggest_category(header)
            idx = CATEGORY_OPTIONS.index(suggested) if suggested in CATEGORY_OPTIONS else 0
            combo.setCurrentIndex(idx)
            table.setCellWidget(row_idx, 1, combo)
            self._combos.append(combo)

            # Sample values.
            sample_vals = [
                row[row_idx] for row in samples
                if row_idx < len(row) and row[row_idx]
            ]
            preview = ", ".join(sample_vals[:3])
            sample_item = QTableWidgetItem(preview)
            sample_item.setFlags(sample_item.flags() & ~sample_item.flags().ItemIsEditable)
            table.setItem(row_idx, 2, sample_item)

        layout.addWidget(table)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def get_mapping(self) -> dict[int, str]:
        """Return ``{col_index: category}`` for every column."""
        return {i: combo.currentText() for i, combo in enumerate(self._combos)}
